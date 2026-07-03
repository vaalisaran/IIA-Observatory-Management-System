from django.contrib import messages
from django.core.paginator import Paginator
from django.shortcuts import get_object_or_404, redirect, render
from django.views import View
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from django.http import HttpResponse

from audit.models import AuditLog
from inventory.models import Branch, BranchStock
from inventory.notifications import notify_inventory_admins
from inventory.utils import (
    get_isolated_products,
    filter_by_branch,
    has_global_inventory_access,
)
from products.models import Product
from ..models import StockEntry


class StockOutPageView(View):
    """
    View to manage outgoing stock reduction transactions.
    """
    def get(self, request):
        # Enforce user authentication
        if not request.user.is_authenticated:
            return redirect("accounts:login")
            
        # Retrieve active stock out log entries filtered by user's branch
        stock_out_entries = filter_by_branch(
            StockEntry.objects.filter(entry_type="out"), request.user
        ).order_by("-timestamp")
        
        # Paginate results list
        paginator = Paginator(stock_out_entries, 50)
        page_obj = paginator.get_page(request.GET.get("page"))
        
        return render(
            request,
            "stock/stock_out.html",
            {
                "stock_out_entries": page_obj.object_list,
                "page_obj": page_obj,
                "products": get_isolated_products(request.user),
                "branches": (
                    Branch.objects.all()
                    if has_global_inventory_access(request.user)
                    else []
                ),
                "current_branch_id": request.GET.get("branch"),
            },
        )

    def post(self, request):
        if not request.user.is_authenticated:
            return redirect("accounts:login")
            
        # Route to bulk processing if requested
        if request.POST.get("form_type") == "bulk":
            return self.handle_bulk_stock_out(request)
            
        # Gather form inputs for manual removal
        product_name, quantity = request.POST.get("product"), int(
            request.POST.get("quantity", 0)
        )
        location_from, location_to, description = (
            request.POST.get("location_from"),
            request.POST.get("location_to"),
            request.POST.get("description"),
        )
        
        # Resolve branch visibility scopes
        branch = (
            get_object_or_404(Branch, id=request.POST.get("branch"))
            if has_global_inventory_access(request.user)
            else getattr(request.user, "branch", None)
        )
        if not branch:
            messages.error(request, "No branch assigned.")
            return redirect("stock-out-page")
            
        # Verify product existence
        product = Product.objects.filter(name__iexact=product_name).first()
        if not product:
            messages.error(request, f'Product "{product_name}" not found.')
            return redirect("stock-out-page")
            
        # Validate that the source branch holds enough quantities to permit subtraction
        source_stock = BranchStock.objects.filter(
            product=product, branch=branch
        ).first()
        available_quantity = source_stock.current_quantity if source_stock else 0
        if quantity > available_quantity:
            messages.error(
                request,
                f"Cannot remove {quantity} units. Only {available_quantity} available in {branch.name}.",
            )
            return redirect("stock-out-page")
            
        # Write stock transaction log entry
        entry = StockEntry.objects.create(
            product=product,
            branch=branch,
            quantity=quantity,
            entry_type="out",
            location_from=location_from,
            location_to=location_to,
            description=description,
            created_by=request.user,
        )
        
        # Record security audit event
        AuditLog.log(request.user, "stock out", entry)
        
        # Alert administrators if triggered by generic workers
        if not request.user.is_admin:
            notify_inventory_admins(
                request.user,
                "stock_out",
                f"Stock Out by {request.user.username}",
                f"{request.user.username} removed {quantity} unit(s) of {product.name} from {branch.name}.",
                target_url="/inventory/stock/out/",
            )
            
        messages.success(
            request,
            f"Successfully removed {quantity} units of {product.name} from {branch.name}.",
        )
        return redirect("stock-out-page")

    def handle_bulk_stock_out(self, request):
        """
        Process multiple stock reduction events parsed from bulk Excel uploads.
        """
        results, success_count, fail_count = [], 0, 0
        if "excel_file" in request.FILES:
            wb = load_workbook(request.FILES["excel_file"])
            ws = wb.active
            header = [cell.value for cell in ws[1]]
            name_idx, qty_idx = header.index("Product Name"), header.index("Quantity")
            
            # Cache products into a dictionary to optimize DB lookup query speeds
            products = {p.name.lower(): p for p in Product.objects.all()}
            for row in ws.iter_rows(min_row=2, values_only=True):
                product_name, qty = str(row[name_idx]).strip(), row[qty_idx]
                product = products.get(product_name.lower())
                
                # Check for validity of row inputs
                if not product or not isinstance(qty, (int, float)) or qty <= 0:
                    results.append(
                        {
                            "product_name": product_name,
                            "quantity": qty,
                            "status": "failed",
                            "message": "Invalid product or quantity",
                        }
                    )
                    fail_count += 1
                    continue
                    
                # Locate appropriate branch scope
                branch_id = request.GET.get("branch") or request.POST.get("branch")
                branch = (
                    Branch.objects.filter(id=branch_id).first()
                    if branch_id
                    else (getattr(request.user, "branch", None) or product.branch)
                )
                if not branch:
                    results.append(
                        {
                            "product_name": product_name,
                            "quantity": qty,
                            "status": "failed",
                            "message": "No branch determined",
                        }
                    )
                    fail_count += 1
                    continue
                    
                # Verify source stock levels
                source_stock = BranchStock.objects.filter(
                    product=product, branch=branch
                ).first()
                available_quantity = (
                    source_stock.current_quantity if source_stock else 0
                )
                if int(qty) > available_quantity:
                    results.append(
                        {
                            "product_name": product.name,
                            "quantity": int(qty),
                            "status": "failed",
                            "message": f"Cannot remove {qty} units. Only {available_quantity} available.",
                        }
                    )
                    fail_count += 1
                    continue
                    
                # Create persistent stock reduction record
                StockEntry.objects.create(
                    product=product,
                    branch=branch,
                    quantity=int(qty),
                    entry_type="out",
                    created_by=request.user,
                )
                results.append(
                    {
                        "product_name": product.name,
                        "quantity": int(qty),
                        "status": "success",
                        "message": "Stock out successful",
                    }
                )
                success_count += 1
                
            # Notify admins of bulk changes
            if success_count and not request.user.is_admin:
                notify_inventory_admins(
                    request.user,
                    "stock_out",
                    f"Bulk Stock Out by {request.user.username}",
                    f"{request.user.username} completed bulk stock-out for {success_count} item(s).",
                    target_url="/inventory/stock/out/",
                )
        if success_count:
            messages.success(
                request, f"Bulk stock out successful for {success_count} item(s)."
            )
        if fail_count:
            messages.error(request, f"Bulk stock out failed for {fail_count} item(s).")
        return render(request, "stock/stock_out.html", {"bulk_results": results})


class DownloadBulkStockOutTemplate(View):
    """
    Constructs and downloads an Excel template for bulk stock-out procedures.
    """
    def get(self, request):
        if not request.user.is_authenticated:
            return redirect("accounts:login")
        wb = Workbook()
        ws = wb.active
        ws.title = "Stock Out Template"
        headers = ["Product Name", "Quantity"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font, cell.fill, cell.alignment = (
                Font(bold=True, color="FFFFFF"),
                PatternFill(
                    start_color="366092", end_color="366092", fill_type="solid"
                ),
                Alignment(horizontal="center"),
            )
        ws.cell(row=2, column=1, value="Sample Product")
        ws.cell(row=2, column=2, value=5)
        ws.column_dimensions["A"].width, ws.column_dimensions["B"].width = 30, 15
        
        # Stream response
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = (
            'attachment; filename="stock_out_template.xlsx"'
        )
        wb.save(response)
        return response
