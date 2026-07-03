from django.contrib import messages
from django.core.paginator import Paginator
from django.shortcuts import get_object_or_404, redirect, render
from django.views import View
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from django.http import HttpResponse

from audit.models import AuditLog
from inventory.models import Branch, BranchStock
from inventory.notifications import notify_inventory_admins
from inventory.utils import (
    get_isolated_products,
    filter_by_branch,
    has_global_inventory_access,
)
from products.models import Product
from ..models import StockEntry


class StockInPageView(View):
    """
    View to display and handle incoming stock entries (single entry or bulk Excel upload).
    Ensures branch isolation and permission verification.
    """
    def get(self, request):
        # Authenticate user session
        if not request.user.is_authenticated:
            return redirect("accounts:login")
            
        # Retrieve incoming stock entries filtered by user's branch permissions
        stock_in_entries = filter_by_branch(
            StockEntry.objects.filter(entry_type="in"), request.user
        ).order_by("-timestamp")
        
        # Paginate results
        paginator = Paginator(stock_in_entries, 50)
        page_obj = paginator.get_page(request.GET.get("page"))
        
        return render(
            request,
            "stock/stock_in.html",
            {
                "stock_in_entries": page_obj.object_list,
                "page_obj": page_obj,
                "products": get_isolated_products(request.user),
                "branches": (
                    Branch.objects.all()
                    if has_global_inventory_access(request.user)
                    else []
                ),
                "current_branch_id": request.GET.get("branch"),
            },
        )

    def post(self, request):
        # Enforce authentication gate
        if not request.user.is_authenticated:
            return redirect("accounts:login")
            
        # Route processing based on upload type
        if request.POST.get("form_type") == "bulk":
            return self.handle_bulk_stock_in(request)
            
        # Extract manual stock entry details
        product_name, quantity = request.POST.get("product"), int(
            request.POST.get("quantity", 0)
        )
        location_from, location_to, description = (
            request.POST.get("location_from"),
            request.POST.get("location_to"),
            request.POST.get("description"),
        )
        
        # Enforce branch check based on user's authorization levels
        branch = (
            get_object_or_404(Branch, id=request.POST.get("branch"))
            if has_global_inventory_access(request.user)
            else getattr(request.user, "branch", None)
        )
        if not branch:
            messages.error(request, "No branch assigned.")
            return redirect("stock-in-page")
            
        # Retrieve target product, matching case-insensitively
        product = Product.objects.filter(name__iexact=product_name).first()
        if not product:
            messages.error(request, f'Product "{product_name}" not found.')
            return redirect("stock-in-page")
            
        # Update physical shelf/rack location mapping inside BranchStock
        rack, shelf = request.POST.get("rack_number"), request.POST.get("shelf_number")
        if rack or shelf:
            bs, _ = BranchStock.objects.get_or_create(product=product, branch=branch)
            if rack:
                bs.rack_number = rack
            if shelf:
                bs.shelf_number = shelf
            bs.save()
            
        # Create persistent database record for the stock transaction
        entry = StockEntry.objects.create(
            product=product,
            branch=branch,
            quantity=quantity,
            entry_type="in",
            location_from=location_from,
            location_to=location_to
            or (f"Rack: {rack}, Shelf: {shelf}" if (rack or shelf) else location_to),
            description=description,
            created_by=request.user,
        )
        
        # Create immutable security audit log entry
        AuditLog.log(request.user, "stock in", entry)
        
        # Dispatch system notification if triggered by non-admin staff
        if not request.user.is_admin:
            notify_inventory_admins(
                request.user,
                "stock_in",
                f"Stock In by {request.user.username}",
                f"{request.user.username} added {quantity} unit(s) of {product.name} to {branch.name}.",
                target_url="/inventory/stock/in/",
            )
            
        messages.success(
            request,
            f"Successfully added {quantity} units of {product.name} at {branch.name}.",
        )
        return redirect("stock-in-page")

    def handle_bulk_stock_in(self, request):
        """
        Parses bulk excel spreadsheet entries to automate multiple stock insertions.
        """
        results, success_count, fail_count = [], 0, 0
        if "excel_file" in request.FILES:
            # Load file object into memory using openpyxl
            wb = load_workbook(request.FILES["excel_file"])
            ws = wb.active
            
            # Map column indices based on header label strings
            header = [cell.value for cell in ws[1]]
            name_idx, qty_idx = header.index("Product Name"), header.index("Quantity")
            rack_idx, shelf_idx = header.index("Rack") if "Rack" in header else None, (
                header.index("Shelf") if "Shelf" in header else None
            )
            
            # Cache products into a dictionary to optimize DB lookup query speeds
            products = {p.name.lower(): p for p in Product.objects.all()}
            
            # Iterate and parse row fields starting from row index 2
            for row in ws.iter_rows(min_row=2, values_only=True):
                product_name, qty = str(row[name_idx]).strip(), row[qty_idx]
                if not product_name or not qty:
                    continue
                product = products.get(product_name.lower())
                if not product:
                    fail_count += 1
                    results.append(
                        {
                            "product_name": product_name,
                            "quantity": qty,
                            "status": "failed",
                            "message": "Product not found",
                        }
                    )
                    continue
                    
                # Determine organizational branch context
                branch_id = request.GET.get("branch") or request.POST.get("branch")
                branch = (
                    Branch.objects.filter(id=branch_id).first()
                    if branch_id
                    else (getattr(request.user, "branch", None) or product.branch)
                )
                if not branch:
                    fail_count += 1
                    results.append(
                        {
                            "product_name": product_name,
                            "quantity": qty,
                            "status": "failed",
                            "message": "No branch determined",
                        }
                    )
                    continue
                    
                rack, shelf = (
                    str(row[rack_idx]).strip()
                    if rack_idx is not None and row[rack_idx]
                    else None
                ), (
                    str(row[shelf_idx]).strip()
                    if shelf_idx is not None and row[shelf_idx]
                    else None
                )
                
                # Update location fields
                if rack or shelf:
                    bs, _ = BranchStock.objects.get_or_create(
                        product=product, branch=branch
                    )
                    if rack:
                        bs.rack_number = rack
                    if shelf:
                        bs.shelf_number = shelf
                    bs.save()
                    
                # Record the validated stock transaction
                StockEntry.objects.create(
                    product=product,
                    branch=branch,
                    quantity=int(qty),
                    entry_type="in",
                    location_to=(
                        f"Rack: {rack}, Shelf: {shelf}"
                        if (rack or shelf)
                        else "Bulk Import"
                    ),
                    created_by=request.user,
                )
                results.append(
                    {
                        "product_name": product.name,
                        "quantity": int(qty),
                        "status": "success",
                        "message": "Stock in successful",
                    }
                )
                success_count += 1
                
            # Dispatch bulk task alerts to admins
            if success_count and not request.user.is_admin:
                notify_inventory_admins(
                    request.user,
                    "stock_in",
                    f"Bulk Stock In by {request.user.username}",
                    f"{request.user.username} completed bulk stock-in for {success_count} item(s).",
                    target_url="/inventory/stock/in/",
                )
        if success_count:
            messages.success(
                request, f"Bulk stock in successful for {success_count} item(s)."
            )
        if fail_count:
            messages.error(request, f"Bulk stock in failed for {fail_count} item(s).")
        return render(request, "stock/stock_in.html", {"bulk_results": results})


class DownloadBulkStockInTemplate(View):
    """
    Constructs and downloads an Excel template for bulk import procedures.
    """
    def get(self, request):
        if not request.user.is_authenticated:
            return redirect("accounts:login")
            
        # Build empty workbook and set layout values
        wb = Workbook()
        ws = wb.active
        ws.title = "Stock In Template"
        headers = ["Product Name", "Quantity"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font, cell.fill, cell.alignment = (
                Font(bold=True, color="FFFFFF"),
                PatternFill(
                    start_color="366092", end_color="366092", fill_type="solid"
                ),
                Alignment(horizontal="center"),
            )
        ws.cell(row=2, column=1, value="Sample Product")
        ws.cell(row=2, column=2, value=10)
        ws.column_dimensions["A"].width, ws.column_dimensions["B"].width = 30, 15
        
        # Stream workbook as response file
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = (
            'attachment; filename="stock_in_template.xlsx"'
        )
        wb.save(response)
        return response
