from django.http import HttpResponse
from django.shortcuts import redirect
from django.views import View
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

"""
This module processes view controllers for downloading Excel upload spreadsheet templates.
"""


class DownloadExcelTemplateView(View):
    """
    Generates and streams an Excel (XLSX) template detailing expected columns
    for bulk product uploads.
    """
    def get(self, request):
        # Enforce authentication constraint to secure the template download
        if not request.user.is_authenticated:
            return redirect("accounts:login")

        # Initialize openpyxl Workbook instance
        wb = Workbook()
        # Fetch the active sheet reference
        ws = wb.active
        # Set sheet title
        ws.title = "Product Template"

        # List of required and optional columns expected by handle_bulk_upload parser
        headers = [
            "Name",
            "Category",
            "Brand",
            "SKU",
            "Model Number",
            "Serial Number",
            "Quantity",
            "Description",
            "Branch (Code)",
            "Local SKU",
            "Rack Number",
            "Shelf Number",
            "Datasheet Filename",
        ]
        
        # Apply premium styling configurations to column headers (bold text, blue background, center aligned)
        for col, header in enumerate(headers, 1):
            # Write header name to the cell in row 1
            cell = ws.cell(row=1, column=col, value=header)
            # Apply styling properties
            cell.font, cell.fill, cell.alignment = (
                Font(bold=True, color="FFFFFF"),
                PatternFill(
                    start_color="366092", end_color="366092", fill_type="solid"
                ),
                Alignment(horizontal="center"),
            )

        # Append a row of sample data to guide users on formatting inputs
        sample_data = [
            "Sample Product",
            "Electronics",
            "Sample Brand",
            "SKU001",
            "MN-990",
            "SN123456",
            "10",
            "Sample product description",
            "IIA",
            "L-SKU001",
            "A1",
            "B2",
            "widget2000.pdf",
        ]
        # Iterate and write sample data to row 2
        for col, value in enumerate(sample_data, 1):
            cell = ws.cell(row=2, column=col, value=value)
            # Use italicized grey font for sample data representation
            cell.font = Font(italic=True, color="666666")

        # Automatically adjust column widths based on maximum text length inside cells
        for column in ws.columns:
            max_length = 0
            for cell in column:
                try:
                    # Calculate string length of cell value
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            # Set the column width to the max length with buffer, capping at 50 units
            ws.column_dimensions[column[0].column_letter].width = min(
                max_length + 2, 50
            )

        # Set standard HTTP response streaming headers for file download (.xlsx format)
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        # Force attachment disposition with the designated filename
        response["Content-Disposition"] = (
            'attachment; filename="product_upload_template.xlsx"'
        )
        # Save workbook contents into the HTTP response stream
        wb.save(response)
        # Return the compiled spreadsheet file
        return response


def download_excel_template(request):
    """
    Functional view wrapper delegating to DownloadExcelTemplateView.
    """
    return DownloadExcelTemplateView().get(request)
