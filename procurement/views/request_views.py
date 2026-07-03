import openpyxl
from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin
from django.db.models import Sum
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.views import View

from inventory.models import BranchStock, InventoryNotification
from inventory.notifications import notify_inventory_admins
from inventory.utils import filter_by_branch, get_isolated_products
from stock.models import StockEntry
from ..models import ProcurementRequest
from products.models import Product

"""
This module processes procurement request creation, imports from spreadsheet lists, and approval/rejections.
"""


class ProcurementUploadView(LoginRequiredMixin, View):
    """
    View class handling rendering lists of procurement requests (GET)
    and submissions of manual or spreadsheet procurement requests (POST).
    Admin decisions (Approval/Rejection) are also processed through this endpoint.
    """
    def get(self, request):
        products = get_isolated_products(request.user)
        status_filter, search = request.GET.get("status", ""), request.GET.get(
            "search", ""
        )
        recent_requests = filter_by_branch(
            ProcurementRequest.objects.select_related("requester", "product", "branch"),
            request.user,
        )
        if status_filter:
            recent_requests = recent_requests.filter(status=status_filter)
        if search:
            recent_requests = recent_requests.filter(product_name__icontains=search)
            
        recent_requests = list(recent_requests.order_by("-created_at")[:100])
        for req in recent_requests:
            if req.product:
                if getattr(request.user, "is_super_admin", False) and not getattr(
                    request.user, "branch", None
                ):
                    # For super admins without branch association, sum stock across all branches
                    req.live_stock = (
                        BranchStock.objects.filter(product=req.product).aggregate(
                            Sum("current_quantity")
                        )["current_quantity__sum"]
                        or 0
                    )
                else:
                    # Otherwise restrict view to the request's target branch or user's assigned branch
                    target_branch = req.branch or getattr(request.user, "branch", None)
                    bs = BranchStock.objects.filter(
                        product=req.product, branch=target_branch
                    ).first()
                    req.live_stock = bs.current_quantity if bs else 0
            else:
                req.live_stock = "-"

        return render(
            request,
            "procurement/upload.html",
            {
                "products": products,
                "recent_requests": recent_requests,
                "status_filter": status_filter,
                "search": search,
            },
        )

    def post(self, request):
        action = request.POST.get("action")
        # Handle admin approval/rejection actions
        if action in ["approve_request", "reject_request"] and request.user.is_admin:
            return self.handle_admin_decision(request, action)
            
        results, insufficient_count, products = (
            [],
            0,
            {p.name.lower(): p for p in Product.objects.all()},
        )

        def process_request(product_name, requested_qty):
            nonlocal insufficient_count
            product = products.get(str(product_name).strip().lower())
            try:
                rq_qty = float(requested_qty)
            except:
                rq_qty = 0
            if not product or rq_qty <= 0:
                results.append(
                    {
                        "product_name": product_name,
                        "requested_qty": requested_qty,
                        "current_stock": "-",
                        "status": "invalid",
                        "product_id": None,
                        "rack_number": "",
                        "shelf_number": "",
                        "alert": False,
                    }
                )
                return
            user_branch = getattr(request.user, "branch", None)
            if not user_branch and request.user.is_super_admin:
                current_stock, rack_number, shelf_number = (
                    BranchStock.objects.filter(product=product).aggregate(
                        Sum("current_quantity")
                    )["current_quantity__sum"]
                    or 0,
                    "Multiple",
                    "Multiple",
                )
            else:
                branch_stock = BranchStock.objects.filter(
                    product=product, branch=user_branch
                ).first()
                current_stock, rack_number, shelf_number = (
                    (branch_stock.current_quantity if branch_stock else 0),
                    (branch_stock.rack_number if branch_stock else "-"),
                    (branch_stock.shelf_number if branch_stock else "-"),
                )
            # Define alert status thresholds
            status, alert = (
                ("out_of_stock", True)
                if current_stock <= 0
                else (("insufficient", True) if current_stock < 10 else ("ok", False))
            )
            if alert:
                insufficient_count += 1
            results.append(
                {
                    "product_name": product.name,
                    "requested_qty": int(rq_qty),
                    "current_stock": current_stock,
                    "status": status,
                    "product_id": product.id,
                    "rack_number": rack_number,
                    "shelf_number": shelf_number,
                    "alert": alert,
                }
            )
            ProcurementRequest.objects.create(
                requester=request.user,
                branch=user_branch,
                product=product,
                product_name=product.name,
                requested_quantity=int(rq_qty),
                current_stock=int(current_stock),
                rack_number=rack_number,
                shelf_number=shelf_number,
                status="pending",
                note="Auto-created from procurement request form",
            )
            if not request.user.is_admin:
                notify_inventory_admins(
                    request.user,
                    "procurement_request",
                    f"Procurement request by {request.user.username}",
                    f"{request.user.username} requested {int(rq_qty)} unit(s) of {product.name}.",
                    target_url="/inventory/procurement/upload/",
                )

        # Process upload via Excel XLSX file
        if "excel_file" in request.FILES:
            try:
                wb = openpyxl.load_workbook(request.FILES["excel_file"])
                ws = wb.active
                header = [cell.value for cell in ws[1]]
                name_idx, qty_idx = header.index("Product Name"), header.index(
                    "Requested Quantity"
                )
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if len(row) > max(name_idx, qty_idx):
                        process_request(row[name_idx], row[qty_idx])
            except Exception as e:
                messages.error(request, f"Error reading spreadsheet file: {str(e)}")
        # Process manual individual request form
        elif request.POST.get("form_type") == "manual":
            process_request(
                request.POST.get("product_name"), request.POST.get("requested_qty")
            )

        if insufficient_count > 0:
            messages.warning(
                request,
                f"There are {insufficient_count} items with insufficient stock.",
            )
        elif results:
            messages.success(request, "Procurement request submitted successfully.")

        # Re-fetch branch products using isolated branch query
        isolated_products = get_isolated_products(request.user)

        # Retrieve and order recent requests exactly like GET logic
        recent_requests = filter_by_branch(
            ProcurementRequest.objects.select_related("requester", "product", "branch"),
            request.user,
        )
        recent_requests = list(recent_requests.order_by("-created_at")[:100])
        for req in recent_requests:
            if req.product:
                if getattr(request.user, "is_super_admin", False) and not getattr(
                    request.user, "branch", None
                ):
                    req.live_stock = (
                        BranchStock.objects.filter(product=req.product).aggregate(
                            Sum("current_quantity")
                        )["current_quantity__sum"]
                        or 0
                    )
                else:
                    target_branch = req.branch or getattr(request.user, "branch", None)
                    bs = BranchStock.objects.filter(
                        product=req.product, branch=target_branch
                    ).first()
                    req.live_stock = bs.current_quantity if bs else 0
            else:
                req.live_stock = "-"

        return render(
            request,
            "procurement/upload.html",
            {
                "results": results,
                "products": isolated_products,
                "recent_requests": recent_requests,
            },
        )

    def handle_admin_decision(self, request, action):
        """
        Processes administrative approvals and rejections of procurement requests.
        Approvals generate a StockEntry (entry_type='in') adding quantities to branch stock levels.
        Rejections notify requester with reason.
        """
        procurement_request = get_object_or_404(
            ProcurementRequest, id=request.POST.get("request_id")
        )
        if procurement_request.status != "pending":
            messages.warning(request, "Request already processed.")
            return redirect("procurement-upload")
        decision_reason = request.POST.get("decision_reason", "").strip()
        
        # Handle Rejections
        if action == "reject_request":
            if not decision_reason:
                messages.error(request, "Rejection reason is required.")
                return redirect("procurement-upload")
            (
                procurement_request.status,
                procurement_request.decision_reason,
                procurement_request.decided_by,
                procurement_request.decided_at,
            ) = ("rejected", decision_reason, request.user, timezone.now())
            procurement_request.save()
            if procurement_request.requester:
                InventoryNotification.objects.create(
                    recipient=procurement_request.requester,
                    sender=request.user,
                    notification_type="procurement_request",
                    title=f"Request #{procurement_request.id} rejected",
                    message=f"Admin rejected request for {procurement_request.product_name}. Reason: {decision_reason}",
                    target_url="/inventory/procurement/upload/",
                )
            messages.success(request, "Procurement request rejected.")
            return redirect("procurement-upload")
            
        # Handle Approvals
        if not procurement_request.product:
            messages.error(request, "Product reference is missing.")
            return redirect("procurement-upload")
        target_branch = (
            procurement_request.branch
            or getattr(procurement_request.requester, "branch", None)
            or getattr(request.user, "branch", None)
        )
        if not target_branch:
            messages.error(request, "No target branch identified.")
            return redirect("procurement-upload")
            
        # Insert a StockEntry transaction (stock goes IN)
        StockEntry.objects.create(
            product=procurement_request.product,
            branch=target_branch,
            quantity=procurement_request.requested_quantity,
            entry_type="in",
            created_by=request.user,
            description=f"Approved Procurement Request #{procurement_request.id}",
        )
        bs = BranchStock.objects.filter(
            product=procurement_request.product, branch=target_branch
        ).first()
        (
            procurement_request.status,
            procurement_request.fulfilled_quantity,
            procurement_request.decision_reason,
            procurement_request.decided_by,
            procurement_request.decided_at,
            procurement_request.current_stock,
        ) = (
            "approved",
            procurement_request.requested_quantity,
            decision_reason or "Approved by admin",
            request.user,
            timezone.now(),
            (bs.current_quantity if bs else 0),
        )
        procurement_request.save()
        if procurement_request.requester:
            InventoryNotification.objects.create(
                recipient=procurement_request.requester,
                sender=request.user,
                notification_type="procurement_request",
                title=f"Request #{procurement_request.id} approved",
                message=f"Admin approved your request for {procurement_request.requested_quantity} unit(s) of {procurement_request.product_name}.",
                target_url="/inventory/procurement/upload/",
            )
        messages.success(
            request,
            f"Request approved. {procurement_request.requested_quantity} units added to {target_branch.name}.",
        )
        return redirect("procurement-upload")
