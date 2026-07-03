import json
import openpyxl
from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import HttpResponse, JsonResponse
from django.shortcuts import redirect
from django.views import View
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST

from inventory.models import Alert
from inventory.notifications import notify_inventory_admins
from products.models import Product

"""
This module processes secondary actions, restock forms, alerts dispatching, and template downloads.
"""


class ProcurementRestockView(LoginRequiredMixin, View):
    """
    View class handling restock submissions for individual items.
    Dispatches notifications to active inventory administrators.
    """
    def post(self, request):
        product_id, requested_qty = request.POST.get("product_id"), request.POST.get(
            "requested_qty"
        ) or request.POST.get("restock_qty")
        if not product_id or not requested_qty:
            messages.error(request, "Invalid product or quantity.")
            return redirect("procurement-upload")
        try:
            product, qty = Product.objects.get(id=product_id), int(requested_qty)
            messages.success(request, f"Restock request for {product.name} submitted!")
            notify_inventory_admins(
                request.user,
                "procurement_request",
                f"Restock action by {request.user.username}",
                f"{request.user.username} submitted restock action for {product.name} ({qty} units).",
                target_url="/inventory/procurement/upload/",
            )
        except Exception:
            messages.error(request, "Error processing restock request.")
        return redirect("procurement-upload")


@csrf_exempt
@require_POST
def send_all_alerts(request):
    """
    API endpoint saving multiple low-stock / out-of-stock alerts in bulk from json payload.
    """
    if not request.user.is_authenticated:
        return JsonResponse({"status": "error", "message": "Authentication required"}, status=401)

    try:
        data = json.loads(request.body.decode("utf-8"))
    except Exception:
        return JsonResponse({"status": "error", "message": "Invalid JSON"}, status=400)

    product_alerts, alert_count = data.get("alerts", []), 0
    for item in product_alerts:
        product_id, requested_qty, current_stock = (
            item.get("product_id"),
            item.get("requested_qty"),
            item.get("current_stock"),
        )
        if not product_id or not requested_qty:
            continue
        try:
            product = Product.objects.get(id=product_id)
            Alert.objects.create(
                product=product,
                alert_type="low_stock" if current_stock > 0 else "out_of_stock",
                status="active",
                message=f"Requested {requested_qty}, but only {current_stock} in stock.",
                current_quantity=current_stock,
                limit_quantity=requested_qty,
            )
            alert_count += 1
            if not request.user.is_admin:
                notify_inventory_admins(
                    request.user,
                    "procurement_request",
                    f"Bulk alerts by {request.user.username}",
                    f"{request.user.username} submitted bulk alerts for {product.name}.",
                    target_url="/inventory/procurement/upload/",
                )
        except Product.DoesNotExist:
            continue
    return JsonResponse({"status": "success", "alert_count": alert_count})


class DownloadProcurementTemplateView(LoginRequiredMixin, View):
    """
    View class generating and streaming a standard template spreadsheet (Excel XLSX format)
    specifying expected columns for bulk imports.
    """
    def get(self, request):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Procurement Template"
        headers = ["Product Name", "Requested Quantity"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font, cell.fill, cell.alignment = (
                Font(bold=True, color="FFFFFF"),
                PatternFill(
                    start_color="366092", end_color="366092", fill_type="solid"
                ),
                Alignment(horizontal="center"),
            )
        ws.cell(row=2, column=1, value="Sample Product")
        ws.cell(row=2, column=2, value=50)
        ws.column_dimensions["A"].width, ws.column_dimensions["B"].width = 30, 20
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = (
            'attachment; filename="procurement_template.xlsx"'
        )
        wb.save(response)
        return response
